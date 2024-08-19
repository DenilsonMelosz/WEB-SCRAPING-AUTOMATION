from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import re
from openpyxl import load_workbook
import os
import logging
import base64
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail, Attachment, FileContent, FileName, FileType, Disposition
from dotenv import load_dotenv

# Carregar variáveis de ambiente do arquivo .env
load_dotenv()

# Obter a chave da API do SendGrid das variáveis de ambiente
SENDGRID_API_KEY = os.getenv('SENDGRID_API_KEY')

# Configuração do logger
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Configure o driver do navegador (usando o Firefox, como no seu exemplo)
options = webdriver.FirefoxOptions()
options.add_argument('--headless')  # Executa o navegador em modo headless (sem interface gráfica)
browser = webdriver.Firefox(options=options)

def extract_data_from_page(url):
    """Extrai dados de uma página específica."""
    data = []
    try:
        browser.get(url)
        
        # Espera explícita para garantir que a lista de produtos esteja carregada
        WebDriverWait(browser, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div[data-testid='product-list']"))
        )
        
        product_list = browser.find_elements(By.CSS_SELECTOR, "div[data-testid='product-list'] li")
        for product in product_list:
            try:
                # Verifique a existência e visibilidade do elemento antes de acessar
                qtd_aval_element = product.find_element(By.CSS_SELECTOR, "span.sc-kUdmhA.dWhxDa")
                QTD_AVAL_TEXT = qtd_aval_element.text

                # Extrai o número de avaliações usando regex
                qtd_aval_match = re.search(r'\((\d+)\)', QTD_AVAL_TEXT)
                if qtd_aval_match:
                    qtd_aval = int(qtd_aval_match.group(1))
                    PRODUTO = product.find_element(By.CSS_SELECTOR, "a div h2").text
                    URL = product.find_element(By.CSS_SELECTOR, "a").get_attribute('href')
                    
                    # Adiciona o item com base na quantidade de avaliações
                    data.append({'Nome': PRODUTO, 'QTD_AVAL': QTD_AVAL_TEXT, 'URL': URL})
                
            except Exception as e:
                logging.error(f"Erro ao processar o produto: {e}")
                continue
    
    except Exception as e:
        logging.error(f"Erro ao acessar a página {url}: {e}")
    
    return data

# Listas para armazenar os dados
melhores_data = []
piores_data = []

# Iterar sobre as páginas de 1 a 17
for page in range(1, 18):
    page_url = f'https://www.magazineluiza.com.br/busca/notebooks/?from=clickSuggestion&page={page}'
    logging.info(f"Coletando dados da página {page}...")
    page_data = extract_data_from_page(page_url)
    for item in page_data:
        # Extrai a quantidade de avaliações do texto
        qtd_aval_match = re.search(r'\((\d+)\)', item['QTD_AVAL'])
        if qtd_aval_match:
            qtd_aval = int(qtd_aval_match.group(1))
            if qtd_aval >= 100:
                melhores_data.append(item)
            else:
                piores_data.append(item)

# Fecha o navegador
browser.quit()

# Convertendo as listas de dados em DataFrames do Pandas
df_melhores = pd.DataFrame(melhores_data)
df_piores = pd.DataFrame(piores_data)

# Renomear a coluna 'Nome' para 'PRODUTO'
df_melhores.rename(columns={'Nome': 'PRODUTO'}, inplace=True)
df_piores.rename(columns={'Nome': 'PRODUTO'}, inplace=True)

# Criar um escritor do Excel para salvar múltiplas abas
excel_path = 'Notebooks.xlsx'
with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    if not df_melhores.empty:
        df_melhores.to_excel(writer, sheet_name='Melhores', index=False)
    if not df_piores.empty:
        df_piores.to_excel(writer, sheet_name='Piores', index=False)

def adjust_column_widths(file_path):
    """Ajusta a largura das colunas no arquivo Excel."""
    if not os.path.isfile(file_path):
        logging.error(f"O arquivo {file_path} não existe.")
        return
    
    try:
        wb = load_workbook(filename=file_path)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except Exception as e:
                        logging.error(f"Erro ao processar a célula: {e}")
                        continue
                adjusted_width = max_length + 2
                ws.column_dimensions[column_letter].width = adjusted_width
        wb.save(filename=file_path)
    except Exception as e:
        logging.error(f"Erro ao ajustar a largura das colunas: {e}")

# Ajustar a largura das colunas no arquivo Excel
adjust_column_widths(excel_path)

logging.info("Olha que legal, a lista de Notebooks, foram recebidas e foi salva em um arquivo Xlxs")

def send_email_via_sendgrid(to_email, subject, body, attachment_path=None):
    """Envia um e-mail usando a API do SendGrid."""
    if not SENDGRID_API_KEY:
        logging.error("SENDGRID_API_KEY não está definida. Verifique seu arquivo .env.")
        return
    
    message = Mail(
        from_email='denilson.testesproc@gmail.com',
        to_emails=to_email,
        subject=subject,
        plain_text_content=body
    )

    # Adicionar anexo, se fornecido
    if attachment_path:
        try:
            with open(attachment_path, 'rb') as f:
                file_data = f.read()
                file_name = os.path.basename(attachment_path)
                encoded_file_data = base64.b64encode(file_data).decode()  # Codificar o arquivo em base64
                attachment = Attachment(
                    FileContent(encoded_file_data),
                    FileName(file_name),
                    FileType('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                    Disposition('attachment')
                )
                message.attachment = attachment
        except Exception as e:
            logging.error(f"Erro ao ler o anexo: {e}")
            return

    try:
        sg = SendGridAPIClient(SENDGRID_API_KEY)
        response = sg.send(message)
        logging.info(f"E-mail enviado! Status code: {response.status_code}")
    except Exception as e:
        logging.error(f"Erro ao enviar e-mail: {str(e)}")

# Exemplo de uso para enviar o e-mail
to_email = 'denis.mello.souza76@gmail.com'
subject = 'Meu Relatório Notebooks'
body = """Olá, aqui está o seu relatório dos notebooks extraídos da Magazine Luiza.

Atenciosamente,
Robô"""
# Defina o caminho do arquivo em seu computador.

attachment_path = r'C:\WEB-SCRAPING-AUTOMATION\Notebooks.xlsx'

send_email_via_sendgrid(to_email, subject, body, attachment_path)
